from __future__ import annotations

import datetime
import logging
import re
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger("vip-parse.bid-comp.xlsx")

_THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

_TITLE_FONT = Font(bold=True, size=16)
_SECTION_HEADER_FONT = Font(bold=True, size=13)
_TABLE_HEADER_FONT = Font(bold=True, size=11)
_FIELD_LABEL_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, size=16, color="FFFFFF")
_SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="D6E4F7")
_SUBHEADER_FONT = Font(size=10, color="1F3864")
_SECTION_HEADER_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
_TABLE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")

_SEVERITY_FILLS = {
    "critical": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    "notable": PatternFill(fill_type="solid", fgColor="FFF3CD"),
    "informational": PatternFill(fill_type="solid", fgColor="D4EDDA"),
}
_MIN_COL_WIDTHS = {
    1: 38,
    2: 16,
    3: 16,
    4: 16,
    5: 16,
    6: 35,
}

_USER_SAFE_FALLBACK_TEXT = "Automated narrative detail unavailable for this section."
_RATE_LIMIT_SAFE_TEXT = (
    "Narrative generation was temporarily rate-limited by the AI provider (HTTP 429). "
    "Core financial comparison results remain available."
)
_INTERNAL_ERROR_PATTERNS = (
    r"\banalysis unavailable\b",
    r"\btoo many requests\b",
    r"\bstatus(?:\s*code)?\s*429\b",
    r"\bhttp(?:x)?\b",
    r"\bopenai\b",
    r"\bchat/completions\b",
    r"\btraceback\b",
    r"\bexception\b",
    r"\bfor more information check\b",
)


def export_xlsx(
    *,
    pair,
    narrative,
    category_rows: List[Dict[str, Any]],
    recap_rows: Optional[List[Dict[str, Any]]] = None,
    methodology: Optional[Any] = None,
    signal_bundle: Optional[Any] = None,
) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_analysis = wb.create_sheet("Analysis")
    report_date = datetime.date.today().strftime("%B %d, %Y")

    _write_summary_sheet(ws_summary, pair, narrative, category_rows, signal_bundle, report_date)
    _write_analysis_sheet(ws_analysis, pair, narrative, category_rows, recap_rows or [], methodology, signal_bundle, report_date)

    bio = BytesIO()
    wb.save(bio)
    payload = bio.getvalue()
    logger.info("xlsx export complete bytes=%d sheets=%d", len(payload), len(wb.sheetnames))
    return payload


def _write_summary_sheet(
    ws,
    pair,
    narrative,
    category_rows: List[Dict[str, Any]],
    signal_bundle: Optional[Any],
    report_date: str,
) -> None:
    _write_sheet_header(
        ws,
        title="Bid Comparison Report",
        pair=pair,
        report_date=report_date,
    )

    primary_total, comparison_total, delta_total = _totals(category_rows)

    ws["A3"] = "Primary Estimate"
    ws["B3"] = pair.primary.estimate_name
    ws["A4"] = "Comparison Estimate"
    ws["B4"] = pair.comparison.estimate_name
    ws["A5"] = "Primary Total"
    ws["B5"] = primary_total
    ws["A6"] = "Comparison Total"
    ws["B6"] = comparison_total
    ws["A7"] = "Total Delta"
    ws["B7"] = delta_total
    for row in range(3, 8):
        ws[f"A{row}"].font = _FIELD_LABEL_FONT
        ws[f"A{row}"].border = _THIN_BORDER
        ws[f"B{row}"].border = _THIN_BORDER
    for row in (5, 6, 7):
        ws[f"B{row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    _style_section_header(ws, 9, "Overall Summary")
    ws["A10"] = _build_overall_summary(narrative)
    ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=6)
    ws["A10"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A10"].border = _THIN_BORDER
    for col in range(2, 7):
        ws.cell(row=10, column=col).border = _THIN_BORDER

    start_row = 12
    _style_section_header(ws, start_row, "Top Cost Drivers")

    headers = ["Category", "Primary ($)", "Comparison ($)", "Delta ($)", "% of Total Variance", "Notes"]
    header_row = start_row + 1
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=header)
        _style_table_header_cell(cell)

    ranked = _ranked_rows(category_rows)
    narrative_map = _narrative_by_category(narrative)
    row = header_row + 1
    for entry in ranked[:6]:
        cat_key = str(entry.get("category") or "").strip().lower()
        driver_text = narrative_map.get(cat_key, "")
        ws.cell(row=row, column=1, value=entry.get("category"))
        ws.cell(row=row, column=2, value=entry.get("primary_total"))
        ws.cell(row=row, column=3, value=entry.get("comparison_total"))
        ws.cell(row=row, column=4, value=entry.get("delta"))
        pct = entry.get("pct_of_total_variance")
        ws.cell(row=row, column=5, value=(pct / 100) if pct is not None else None)
        ws.cell(row=row, column=6, value=driver_text or None)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = _THIN_BORDER
        for col in (2, 3, 4):
            ws.cell(row=row, column=col).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=row, column=5).number_format = "0.0%"
        if driver_text:
            ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=6).font = Font(italic=True, size=9, color="444444")
        row += 1

    observations_row = row + 1
    _style_section_header(ws, observations_row, "Key Observations")
    observations_row += 1
    observations = _build_observations(narrative, signal_bundle)
    for item in observations:
        ws.cell(row=observations_row, column=1, value=f"- {item}")
        ws.merge_cells(start_row=observations_row, start_column=1, end_row=observations_row, end_column=6)
        ws.cell(row=observations_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        observations_row += 1

    _configure_print(ws)
    ws.freeze_panes = "A3"
    _autosize(ws)


def _write_analysis_sheet(
    ws,
    pair,
    narrative,
    category_rows: List[Dict[str, Any]],
    recap_rows: List[Dict[str, Any]],
    methodology: Optional[Any],
    signal_bundle: Optional[Any],
    report_date: str,
) -> None:
    _write_sheet_header(
        ws,
        title="Bid Comparison - Analysis",
        pair=pair,
        report_date=report_date,
    )
    row = 3

    _style_section_header(ws, row, "Methodology Detail")
    row += 1
    method_rows = _methodology_rows(methodology)
    for label, value in method_rows:
        ws.cell(row=row, column=1, value=label).font = _FIELD_LABEL_FONT
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).border = _THIN_BORDER
        ws.cell(row=row, column=2).border = _THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    row += 1
    _style_section_header(ws, row, "Scope Alignment")
    row += 1
    scope_items = list(getattr(narrative, "scope_observations", []) or [])
    scope_items = [item for item in (_sanitize_user_text(item) for item in scope_items) if item]
    if not scope_items:
        scope_items = ["No scope observations"]
    for item in scope_items:
        ws.cell(row=row, column=1, value=f"- {item}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    followups = list(getattr(narrative, "suggested_followups", []) or [])
    followups = [item for item in (_sanitize_user_text(item) for item in followups) if item]
    if followups:
        row += 1
        _style_section_header(ws, row, "Suggested Follow-ups")
        row += 1
        for item in followups:
            ws.cell(row=row, column=1, value=f"- {item}")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    row += 1
    _style_section_header(ws, row, "Category-by-Category Comparison")
    row += 1
    headers = [
        "Category",
        f"{pair.primary.estimate_name} ($)",
        f"{pair.comparison.estimate_name} ($)",
        "Delta ($)",
        "Delta (% of Primary)",
        "Severity",
    ]
    table_header_row = row
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=table_header_row, column=idx, value=header)
        _style_table_header_cell(cell)

    severity_by_category = _severity_by_category(signal_bundle)
    row = table_header_row + 1
    for item in category_rows:
        category = item.get("category")
        severity = severity_by_category.get(str(category or ""), "")
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=2, value=item.get("primary_total"))
        ws.cell(row=row, column=3, value=item.get("comparison_total"))
        ws.cell(row=row, column=4, value=item.get("delta"))
        ws.cell(row=row, column=5, value=item.get("delta_pct"))
        ws.cell(row=row, column=6, value=severity)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = _THIN_BORDER
        for col in (2, 3, 4):
            ws.cell(row=row, column=col).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws.cell(row=row, column=5).number_format = "0.00%"
        fill = _SEVERITY_FILLS.get(str(severity).lower())
        if fill is not None:
            ws.cell(row=row, column=6).fill = fill
        row += 1

    table_end_row = row - 1
    if table_end_row >= table_header_row + 1:
        ws.conditional_formatting.add(
            f"D{table_header_row + 1}:D{table_end_row}",
            ColorScaleRule(
                start_type="num",
                start_value=-1,
                start_color="F4CCCC",
                mid_type="num",
                mid_value=0,
                mid_color="FFFFFF",
                end_type="num",
                end_value=1,
                end_color="D4EDDA",
            ),
        )
        ws.conditional_formatting.add(
            f"D{table_header_row + 1}:D{table_end_row}",
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="max",
                end_value=0,
                color="638EC6",
                showValue=True,
            ),
        )

    if recap_rows:
        row += 2
        _style_section_header(ws, row, "Recap Detail (raw)")
        row += 1
        recap_headers = ["Estimate", "Group", "Item", "Total ($)"]
        for idx, header in enumerate(recap_headers, start=1):
            cell = ws.cell(row=row, column=idx, value=header)
            _style_table_header_cell(cell)
        row += 1
        for entry in recap_rows[:500]:
            ws.cell(row=row, column=1, value=entry.get("estimate"))
            ws.cell(row=row, column=2, value=entry.get("group"))
            ws.cell(row=row, column=3, value=entry.get("item"))
            ws.cell(row=row, column=4, value=entry.get("total"))
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = _THIN_BORDER
            ws.cell(row=row, column=4).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            row += 1

    _configure_print(ws)
    ws.freeze_panes = "A3"
    _autosize(ws)


def _build_observations(narrative, signal_bundle: Optional[Any]) -> List[str]:
    observations: List[str] = []

    overview = _sanitize_user_text(getattr(narrative, "overview", ""))
    if overview:
        observations.append(overview)

    key_drivers = list(getattr(narrative, "key_drivers", []) or [])
    for item in key_drivers[:3]:
        if hasattr(item, "narrative"):
            text = str(getattr(item, "narrative") or "").strip()
        else:
            text = str((item or {}).get("narrative") or "").strip()
        sanitized = _sanitize_user_text(text)
        if sanitized:
            observations.append(sanitized)

    if signal_bundle is not None:
        for alert in getattr(signal_bundle, "alert_tags", []) or []:
            severity = getattr(getattr(alert, "severity", None), "value", getattr(alert, "severity", ""))
            title = getattr(alert, "title", "")
            detail = getattr(alert, "detail", "")
            sanitized = _sanitize_user_text(f"[{severity}] {title}: {detail}".strip())
            if sanitized:
                observations.append(sanitized)

    if not observations:
        return ["No key observations available"]
    return observations


def _build_overall_summary(narrative) -> str:
    overview = _sanitize_user_text(getattr(narrative, "overview", ""))
    if overview:
        return overview
    return "Comparison summary unavailable. Refer to top cost drivers and analysis details below."


def _methodology_rows(methodology: Optional[Any]) -> List[tuple[str, str]]:
    if methodology is None:
        return [("Summary", "Methodology analysis not available")]

    return [
        (
            "Summary",
            (
                f"O&P: {methodology.primary_op.structure_type.value} vs {methodology.comparison_op.structure_type.value}; "
                f"Depreciation differs: {methodology.depreciation_approach_differs}; "
                f"Price list differs: {methodology.price_list_differs}; "
                f"Granularity: {methodology.data_granularity.value}"
            ),
        ),
        ("Primary O&P", str(methodology.primary_op.structure_type.value)),
        ("Comparison O&P", str(methodology.comparison_op.structure_type.value)),
        ("Depreciation Differs", str(methodology.depreciation_approach_differs)),
        ("Price List Differs", str(methodology.price_list_differs)),
        ("Locality Factors", str(methodology.locality_factors or "n/a")),
        ("Data Granularity", str(methodology.data_granularity.value)),
    ]


def _ranked_rows(category_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    ranked: List[Dict[str, Any]] = []
    total_abs = sum(abs(float(r.get("delta") or 0.0)) for r in category_rows) or 1.0
    for row in category_rows:
        delta = float(row.get("delta") or 0.0)
        ranked.append(
            {
                "category": row.get("category") or "Unknown",
                "primary_total": float(row.get("primary_total") or 0.0),
                "comparison_total": float(row.get("comparison_total") or 0.0),
                "delta": delta,
                "abs_delta": abs(delta),
                "pct_of_total_variance": round(abs(delta) / total_abs * 100.0, 2),
            }
        )
    ranked.sort(key=lambda r: (-r["abs_delta"], str(r["category"]).lower()))
    return ranked


def _totals(category_rows: List[Dict[str, Any]]) -> tuple[float, float, float]:
    primary_total = round(sum(float(r.get("primary_total") or 0.0) for r in category_rows), 2)
    comparison_total = round(sum(float(r.get("comparison_total") or 0.0) for r in category_rows), 2)
    delta_total = round(comparison_total - primary_total, 2)
    return primary_total, comparison_total, delta_total


def _severity_by_category(signal_bundle: Optional[Any]) -> Dict[str, str]:
    if signal_bundle is None:
        return {}
    out: Dict[str, str] = {}
    for flag in getattr(signal_bundle, "emphasis_flags", []) or []:
        category = getattr(flag, "category", None)
        severity = getattr(getattr(flag, "severity", None), "value", None)
        if category and severity:
            out[str(category)] = str(severity)
    return out


def _narrative_by_category(narrative: Optional[Any]) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for item in getattr(narrative, "key_drivers", []) or []:
        if isinstance(item, dict):
            cat = str(item.get("category") or "").strip().lower()
            text = _sanitize_user_text(str(item.get("narrative") or "").strip())
        elif hasattr(item, "narrative"):
            cat = str(getattr(item, "category", "")).strip().lower()
            text = _sanitize_user_text(str(getattr(item, "narrative", "")).strip())
        else:
            continue
        if cat and text:
            result[cat] = text
    return result


def _write_sheet_header(ws, title: str, pair, report_date: str) -> None:
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = _HEADER_FONT
    ws["A1"].fill = _HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    for col in range(2, 7):
        ws.cell(row=1, column=col).fill = _HEADER_FILL

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Report Date: {report_date}  |  "
        f"Primary: {pair.primary.estimate_name}  |  "
        f"Comparison: {pair.comparison.estimate_name}"
    )
    ws["A2"].font = _SUBHEADER_FONT
    ws["A2"].fill = _SUBHEADER_FILL
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18
    for col in range(2, 7):
        ws.cell(row=2, column=col).fill = _SUBHEADER_FILL


def _style_section_header(ws, row: int, title: str) -> None:
    ws.cell(row=row, column=1, value=title).font = _SECTION_HEADER_FONT
    ws.cell(row=row, column=1).fill = _SECTION_HEADER_FILL
    ws.cell(row=row, column=1).border = _THIN_BORDER
    for col in range(2, 7):
        ws.cell(row=row, column=col).fill = _SECTION_HEADER_FILL
        ws.cell(row=row, column=col).border = _THIN_BORDER


def _style_table_header_cell(cell) -> None:
    cell.font = _TABLE_HEADER_FONT
    cell.border = _THIN_BORDER
    cell.fill = _TABLE_HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _configure_print(ws) -> None:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    ws.print_options.horizontalCentered = True


def _autosize(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (
                len(str(cell.value))
                for cell in column_cells
                if cell.value is not None and not isinstance(cell, MergedCell)
            ),
            default=0,
        )
        min_width = _MIN_COL_WIDTHS.get(col_idx, 12)
        ws.column_dimensions[letter].width = min(max(min_width, max_len + 3), 90)


def _sanitize_user_text(value: Any) -> Optional[str]:
    text = str(value or "").strip()
    if not text:
        return None
    lowered = text.lower()
    if re.search(r"\btoo many requests\b|\bstatus(?:\s*code)?\s*429\b", lowered):
        return _RATE_LIMIT_SAFE_TEXT
    for pattern in _INTERNAL_ERROR_PATTERNS:
        if re.search(pattern, lowered):
            return _USER_SAFE_FALLBACK_TEXT
    return text
