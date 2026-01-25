from __future__ import annotations

import logging
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .markdown import MarkdownBlock

logger = logging.getLogger("vip-parse.bid-comp.xlsx")

def export_xlsx(
    *,
    pair,
    narrative,
    category_rows: List[Dict[str, Any]],
    recap_rows: List[Dict[str, Any]],
) -> bytes:
    logger.info(
        "xlsx export start: primary=%s comparison=%s category_rows=%d recap_rows=%d",
        pair.primary.estimate_name,
        pair.comparison.estimate_name,
        len(category_rows),
        len(recap_rows),
    )
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Narrative Summary"
    header_font = Font(bold=True)

    ws_summary["A1"] = "Bid Comparison Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)

    ws_summary["A3"] = "Estimate"
    ws_summary["B3"] = "Grand Total ($)"

    # Get totals with fallback computation from category_rows if missing
    primary_total = pair.primary.totals.grand_total
    comparison_total = pair.comparison.totals.grand_total

    # Log what we have
    logger.info(
        "xlsx totals: primary=%s (%s) comparison=%s (%s)",
        primary_total,
        type(primary_total).__name__,
        comparison_total,
        type(comparison_total).__name__,
    )

    # Fallback: compute from category_rows if grand_total is missing
    if primary_total is None and category_rows:
        primary_total = sum(
            (r.get("primary_total") or 0) for r in category_rows
            if isinstance(r.get("primary_total"), (int, float))
        )
        logger.warning("xlsx primary grand_total missing, computed from categories: %s", primary_total)

    if comparison_total is None and category_rows:
        comparison_total = sum(
            (r.get("comparison_total") or 0) for r in category_rows
            if isinstance(r.get("comparison_total"), (int, float))
        )
        logger.warning("xlsx comparison grand_total missing, computed from categories: %s", comparison_total)

    # Ensure we have numbers, not None
    primary_total = primary_total or 0
    comparison_total = comparison_total or 0
    delta_total = comparison_total - primary_total

    ws_summary["A4"] = pair.primary.estimate_name
    ws_summary["B4"] = primary_total
    ws_summary["A5"] = pair.comparison.estimate_name
    ws_summary["B5"] = comparison_total

    # Add bold delta row
    ws_summary["A6"] = "Delta"
    ws_summary["A6"].font = Font(bold=True)
    ws_summary["B6"] = delta_total
    ws_summary["B6"].font = Font(bold=True)

    for cell_ref in ("B4", "B5", "B6"):
        cell = ws_summary[cell_ref]
        if isinstance(cell.value, (int, float)):
            cell.number_format = "$#,##0.00"

    sections = narrative.sections or {}
    overview_text = _resolve_overview_text(narrative)
    scope_items = _ensure_string_list(sections.get("scope_observations"))
    followup_items = _ensure_string_list(sections.get("suggested_followups"))
    driver_rows = _ensure_driver_rows(sections.get("key_cost_drivers"), narrative, pair)
    driver_source = (
        "sections"
        if sections.get("key_cost_drivers")
        else "narrative"
        if getattr(narrative, "key_drivers", None)
        else "delta_rows"
    )
    sample_preview = ""
    for entry in driver_rows:
        text = (entry.get("narrative") or "").strip()
        if text:
            sample_preview = text[:120]
            break
    if driver_rows and not sample_preview:
        logger.warning("xlsx driver rows missing narrative text despite %d entries", len(driver_rows))
    logger.info(
        "xlsx driver rows: source=%s count=%d sample=%s",
        driver_source,
        len(driver_rows),
        sample_preview or "<empty>",
    )

    current_row = 8  # After header (1), blank (2), estimate header (3), primary (4), comparison (5), delta (6), blank (7)
    try:
        current_row = _write_text_section(ws_summary, current_row, "Overview of Estimates", overview_text, header_font)
    except Exception:  # noqa: BLE001
        logger.exception("xlsx summary section write failed: overview")
        raise
    try:
        current_row = _write_key_driver_section(
            ws_summary,
            current_row,
            pair,
            driver_rows,
            header_font,
        )
    except Exception:  # noqa: BLE001
        logger.exception("xlsx summary section write failed: key cost drivers")
        raise
    try:
        current_row = _write_list_section(ws_summary, current_row, "Scope Observations", scope_items, header_font)
    except Exception:  # noqa: BLE001
        logger.exception("xlsx summary section write failed: scope observations")
        raise
    try:
        current_row = _write_list_section(ws_summary, current_row, "Suggested Follow-ups", followup_items, header_font)
    except Exception:  # noqa: BLE001
        logger.exception("xlsx summary section write failed: suggested follow-ups")
        raise

    if narrative.delta_rows:
        current_row += 1
        ws_summary.cell(row=current_row, column=1, value="Key Category Deltas").font = header_font
        current_row += 1
        delta_headers = [
            "Category",
            f"{pair.primary.estimate_name} ($)",
            f"{pair.comparison.estimate_name} ($)",
            "Delta ($)",
        ]
        for col_idx, header in enumerate(delta_headers, start=1):
            ws_summary.cell(row=current_row, column=col_idx, value=header).font = header_font
        current_row += 1
        for entry in narrative.delta_rows:
            ws_summary.cell(row=current_row, column=1, value=entry.get("category"))
            ws_summary.cell(row=current_row, column=2, value=entry.get("primary_total"))
            ws_summary.cell(row=current_row, column=3, value=entry.get("comparison_total"))
            ws_summary.cell(row=current_row, column=4, value=entry.get("delta"))
            current_row += 1
        for col_idx in (2, 3, 4):
            col_letter = get_column_letter(col_idx)
            for cell in ws_summary[col_letter]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "$#,##0.00"

    _autosize(ws_summary)
    logger.info(
        "xlsx summary sheet complete: primary_total=%s comparison_total=%s rows=%d",
        pair.primary.totals.grand_total,
        pair.comparison.totals.grand_total,
        current_row,
    )

    # Sheet 2: category matrix
    ws_categories = wb.create_sheet("Verisk Categories")
    cat_headers = [
        "Category",
        f"{pair.primary.estimate_name} ($)",
        f"{pair.comparison.estimate_name} ($)",
        "Delta ($)",
        f"Delta (% of {pair.primary.estimate_name})",
    ]
    ws_categories.append(cat_headers)
    for col_idx in range(1, len(cat_headers) + 1):
        ws_categories.cell(row=1, column=col_idx).font = header_font
    for row in category_rows:
        ws_categories.append(
            [
                row["category"],
                row.get("primary_total"),
                row.get("comparison_total"),
                row.get("delta"),
                row.get("delta_pct"),
            ]
        )
    for col_idx in (2, 3, 4):
        col_letter = get_column_letter(col_idx)
        for cell in ws_categories[col_letter]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "$#,##0.00"
    pct_column = get_column_letter(5)
    for cell in ws_categories[pct_column]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00%"
    ws_categories.freeze_panes = "A2"
    _autosize(ws_categories)
    logger.info("xlsx category sheet complete: rows=%d", len(category_rows))

    # Sheet 3: raw recap
    ws_recap = wb.create_sheet("Original Recap")
    recap_headers = ["Estimate", "Group", "Item", "Total ($)"]
    ws_recap.append(recap_headers)
    for col_idx in range(1, len(recap_headers) + 1):
        ws_recap.cell(row=1, column=col_idx).font = header_font
    for entry in recap_rows:
        ws_recap.append(
            [
                entry.get("estimate"),
                entry.get("group"),
                entry.get("item"),
                entry.get("total"),
            ]
        )
    total_col = get_column_letter(4)
    for cell in ws_recap[total_col]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = "$#,##0.00"
    ws_recap.freeze_panes = "A2"
    _autosize(ws_recap)
    logger.info("xlsx recap sheet complete: rows=%d", len(recap_rows))

    bio = BytesIO()
    wb.save(bio)
    data = bio.getvalue()
    logger.info("xlsx export complete: bytes=%d sheets=%d", len(data), len(wb.sheetnames))
    return data


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        max_len = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 80)

def _write_text_section(ws, start_row: int, title: str, body: str, header_font: Font) -> int:
    ws.cell(row=start_row, column=1, value=title).font = header_font
    body_row = start_row + 1
    resolved_body = body or "No narrative available."
    if not body:
        logger.warning("xlsx narrative text missing for section '%s'; using fallback", title)

    # Split on double newlines first (paragraph breaks), then single newlines
    # This preserves intentional paragraph structure from LLM output
    paragraphs = resolved_body.split('\n\n')
    structured_lines = []

    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        # Check if this paragraph starts with a bold label
        if para.startswith('**'):
            # Extract label and value from **Label**: Value format
            if '**:' in para:
                parts = para.split('**:', 1)
                label = parts[0].replace('**', '').strip()
                value = parts[1].strip() if len(parts) > 1 else ''
                structured_lines.append((label, value))
            elif para.count('**') >= 2:
                # Format: **Label** Value or **Label**: Value
                end_bold = para.find('**', 2)
                if end_bold > 0:
                    label = para[2:end_bold].strip()
                    rest = para[end_bold+2:].lstrip(':').strip()
                    structured_lines.append((label, rest))
                else:
                    structured_lines.append((None, para))
            else:
                structured_lines.append((None, para))
        elif para.startswith('- **'):
            # Bullet with bold label
            para = para.lstrip('- ')
            if '**:' in para:
                parts = para.split('**:', 1)
                label = parts[0].replace('**', '').strip()
                value = parts[1].strip() if len(parts) > 1 else ''
                structured_lines.append((label, value))
            else:
                structured_lines.append((None, para))
        else:
            structured_lines.append((None, para))

    if structured_lines and any(label for label, _ in structured_lines):
        # Write as structured rows with bold labels
        for label, value in structured_lines:
            if label:
                label_cell = ws.cell(row=body_row, column=1, value=f"{label}:")
                label_cell.font = Font(bold=True)
                ws.merge_cells(start_row=body_row, start_column=2, end_row=body_row, end_column=5)
                value_cell = ws.cell(row=body_row, column=2, value=value)
                value_cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                ws.merge_cells(start_row=body_row, start_column=1, end_row=body_row, end_column=5)
                text_cell = ws.cell(row=body_row, column=1, value=value)
                text_cell.alignment = Alignment(wrap_text=True, vertical="top")
            body_row += 1
        return body_row + 1
    else:
        # Fallback to single merged cell for plain text
        ws.merge_cells(start_row=body_row, start_column=1, end_row=body_row, end_column=5)
        body_cell = ws.cell(row=body_row, column=1, value=resolved_body)
        body_cell.alignment = Alignment(wrap_text=True, vertical="top")
        return body_row + 2


def _write_list_section(ws, start_row: int, title: str, items: List[str], header_font: Font) -> int:
    ws.cell(row=start_row, column=1, value=title).font = header_font
    row = start_row + 1
    if not items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value="No items provided.")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        return row + 2
    for item in items:
        ws.cell(row=row, column=1, value="•")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        text_cell = ws.cell(row=row, column=2, value=item)
        text_cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    return row + 1


def _write_key_driver_section(ws, start_row: int, pair, drivers: List[Dict[str, Any]], header_font: Font) -> int:
    ws.cell(row=start_row, column=1, value="Key Cost Drivers").font = header_font
    row = start_row + 1
    headers = [
        "Category",
        f"{pair.primary.estimate_name} ($)",
        f"{pair.comparison.estimate_name} ($)",
        "Delta ($)",
        "Narrative",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=header).font = header_font
    row += 1
    if not drivers:
        logger.warning(
            "xlsx driver rows empty: primary=%s comparison=%s",
            pair.primary.estimate_name,
            pair.comparison.estimate_name,
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value="No driver data provided.")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        return row + 2
    # Track totals for summary row
    total_primary = 0.0
    total_comparison = 0.0
    total_delta = 0.0

    for driver in drivers:
        primary_val = driver.get("primary_total")
        comparison_val = driver.get("comparison_total")
        delta_val = driver.get("delta_total")
        if delta_val is None and isinstance(primary_val, (int, float)) and isinstance(comparison_val, (int, float)):
            delta_val = round((comparison_val or 0) - (primary_val or 0), 2)

        # Accumulate totals
        if isinstance(primary_val, (int, float)):
            total_primary += primary_val
        if isinstance(comparison_val, (int, float)):
            total_comparison += comparison_val
        if isinstance(delta_val, (int, float)):
            total_delta += delta_val

        ws.cell(row=row, column=1, value=driver.get("category"))
        primary_cell = ws.cell(row=row, column=2, value=primary_val)
        comparison_cell = ws.cell(row=row, column=3, value=comparison_val)
        delta_cell = ws.cell(row=row, column=4, value=delta_val)
        narrative_cell = ws.cell(row=row, column=5, value=driver.get("narrative") or "")
        for cell in (primary_cell, comparison_cell, delta_cell):
            if isinstance(cell.value, (int, float)):
                cell.number_format = "$#,##0.00"
        narrative_cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # Add Grand Total row
    ws.cell(row=row, column=1, value="Grand Total").font = header_font
    total_primary_cell = ws.cell(row=row, column=2, value=total_primary)
    total_primary_cell.font = header_font
    total_primary_cell.number_format = "$#,##0.00"
    total_comparison_cell = ws.cell(row=row, column=3, value=total_comparison)
    total_comparison_cell.font = header_font
    total_comparison_cell.number_format = "$#,##0.00"
    total_delta_cell = ws.cell(row=row, column=4, value=total_delta)
    total_delta_cell.font = header_font
    total_delta_cell.number_format = "$#,##0.00"
    row += 1

    return row + 1


def _ensure_string_list(value) -> List[str]:
    if not isinstance(value, list):
        return []
    items: List[str] = []
    for entry in value:
        if isinstance(entry, str):
            text = entry.strip()
            if text:
                items.append(text)
    return items


def _ensure_driver_rows(section_rows, narrative, pair=None) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    fallback_count = 0
    llm_count = 0
    if isinstance(section_rows, list):
        for entry in section_rows:
            normalized_entry = _normalize_driver_entry(entry)
            if normalized_entry:
                normalized.append(normalized_entry)
        if section_rows and not normalized:
            logger.warning("xlsx driver rows dropped after normalization: %d entries", len(section_rows))
    if normalized:
        # Ensure all drivers have narratives, generate fallback if missing
        for driver in normalized:
            if not driver.get("narrative"):
                driver["narrative"] = _generate_fallback_narrative(driver, pair)
                fallback_count += 1
            else:
                llm_count += 1
        logger.info("xlsx driver narratives: llm=%d fallback=%d", llm_count, fallback_count)
        return normalized
    if getattr(narrative, "key_drivers", None):
        fallback_rows = [_normalize_driver_entry(entry) or entry for entry in narrative.key_drivers]
        if not fallback_rows:
            logger.warning("xlsx driver fallback (narrative.key_drivers) produced 0 entries")
        # Ensure all drivers have narratives
        for driver in fallback_rows:
            if not driver.get("narrative"):
                driver["narrative"] = _generate_fallback_narrative(driver, pair)
                fallback_count += 1
            else:
                llm_count += 1
        logger.info("xlsx driver narratives (key_drivers path): llm=%d fallback=%d", llm_count, fallback_count)
        return fallback_rows
    # Full fallback path - no LLM narratives at all
    drivers: List[Dict[str, Any]] = []
    for row in (narrative.delta_rows or []):
        driver = {
            "category": row.get("category"),
            "primary_total": row.get("primary_total"),
            "comparison_total": row.get("comparison_total"),
            "delta_total": row.get("delta"),
            "narrative": "",
        }
        driver["narrative"] = _generate_fallback_narrative(driver, pair)
        drivers.append(driver)
    if drivers:
        logger.warning("xlsx driver narratives: ALL FALLBACK (%d drivers) - LLM returned no narratives", len(drivers))
    else:
        logger.warning("xlsx driver fallback (delta_rows) produced 0 entries")
    return drivers


def _generate_fallback_narrative(driver: Dict[str, Any], pair=None) -> str:
    """Generate a basic narrative when LLM doesn't provide one."""
    category = driver.get("category", "Category")
    primary_total = driver.get("primary_total")
    comparison_total = driver.get("comparison_total")
    delta = driver.get("delta_total")

    if delta is None and primary_total is not None and comparison_total is not None:
        delta = comparison_total - primary_total

    if delta is None:
        return f"Review {category} line items for scope differences."

    primary_name = pair.primary.estimate_name if pair else "Primary"
    comparison_name = pair.comparison.estimate_name if pair else "Comparison"

    abs_delta = abs(delta)
    if abs_delta < 100:
        return f"Minor variance in {category}; likely rounding or unit price differences."

    if delta > 0:
        # Comparison is higher
        pct = (delta / primary_total * 100) if primary_total and primary_total > 0 else 0
        if pct > 50:
            return f"{comparison_name} includes significantly more {category} scope (+{pct:.0f}%); review for additional line items or quantities."
        elif pct > 20:
            return f"{comparison_name} higher in {category} (+{pct:.0f}%); check for quantity or unit price differences."
        else:
            return f"Moderate increase in {category} for {comparison_name}; may reflect scope additions or pricing variance."
    else:
        # Primary is higher
        pct = (abs_delta / primary_total * 100) if primary_total and primary_total > 0 else 0
        if pct > 50:
            return f"{comparison_name} missing or reduced {category} scope (-{pct:.0f}%); verify if items were excluded."
        elif pct > 20:
            return f"{comparison_name} lower in {category} (-{pct:.0f}%); check for missing line items or reduced quantities."
        else:
            return f"Moderate decrease in {category} for {comparison_name}; may reflect scope reductions or pricing variance."


def _resolve_overview_text(narrative) -> str:
    sections = narrative.sections or {}
    overview = (sections.get("overview_of_estimates") or "").strip()
    if overview:
        return overview
    overview_text = _extract_first_paragraph(narrative.blocks)
    if not overview_text:
        logger.warning("narrative overview missing: falling back to default")
    return overview_text or "No narrative available."


def _extract_first_paragraph(blocks: List[MarkdownBlock]) -> str:
    first_heading: Optional[str] = None
    for block in blocks:
        if block.kind == "heading" and block.text and not first_heading:
            first_heading = block.text
            continue
        if block.kind == "paragraph" and block.text:
            return block.text
    return first_heading or ""


def _normalize_driver_entry(entry: Any) -> Dict[str, Any] | None:
    if not isinstance(entry, dict):
        return None

    def _pick_number(*keys):
        for key in keys:
            if key in entry:
                value = entry.get(key)
                coerced = _coerce_number(value)
                if coerced is not None:
                    return coerced
        return None

    category = str(entry.get("category") or "").strip() or "Category"
    primary_total = _pick_number("total1", "primary_total", "estimate_a_total", "left_total")
    comparison_total = _pick_number("total2", "comparison_total", "estimate_b_total", "right_total")
    delta_val = _pick_number("delta", "delta_total")
    if delta_val is None and isinstance(primary_total, (int, float)) and isinstance(comparison_total, (int, float)):
        delta_val = round(comparison_total - primary_total, 2)
    narrative_text = str(entry.get("narrative") or entry.get("explanation") or "").strip()
    return {
        "category": category,
        "primary_total": primary_total,
        "comparison_total": comparison_total,
        "delta_total": delta_val,
        "narrative": narrative_text,
    }


def _coerce_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
