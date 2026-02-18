from __future__ import annotations

import json
from io import BytesIO

from openpyxl import load_workbook

from vip_shared.bid_comp import BidComp
from vip_shared.llm.adapter import LLMAdapterBase


def _make_payload(name: str, framing: float, roofing: float, electrical: float, overhead: float, profit: float, tax: float) -> dict:
    recap = {
        "O&P Items": [
            {"item": "FRAMING", "total": f"{framing:,.2f}"},
            {"item": "ROOFING", "total": f"{roofing:,.2f}"},
        ],
        "Non-O&P Items": [
            {"item": "ELECTRICAL", "total": f"{electrical:,.2f}"},
        ],
        "subtotals": [
            {"label": "Overhead", "total": f"{overhead:,.2f}"},
            {"label": "Profit", "total": f"{profit:,.2f}"},
            {"label": "Material Sales Tax", "total": f"{tax:,.2f}"},
            {"label": "Total", "total": f"{framing + roofing + electrical + overhead + profit + tax:,.2f}"},
        ],
    }
    return {
        "estimate_name": name,
        "case_metadata": {
            "estimate_name": name,
            "line_item_totals": {
                "grand_total": f"{framing + roofing + electrical + overhead + profit + tax:,.2f}",
                "material_sales_tax": f"{tax:,.2f}",
                "overhead_profit": f"{overhead + profit:,.2f}",
            },
        },
        "recaps_and_summaries": {"recap_by_category": recap},
        "sections": [],
    }


class FakeAdapter(LLMAdapterBase):
    def __init__(self, markdown: str, sections: dict | None = None) -> None:
        payload = {"markdown": markdown, "metadata": {"confidence": "high"}}
        if sections is not None:
            payload["sections"] = sections
        self.response = json.dumps(payload)

    def generate(self, template_id: str, context: dict) -> str:  # type: ignore[override]
        assert template_id == "bid_comp_summary_v1"
        assert "primary_json" in context
        assert "comparison_json" in context
        assert "category_table_json" in context
        return self.response


def test_category_mapping_and_fallback_narrative() -> None:
    payload_a = _make_payload("Estimate A", framing=400, roofing=100, electrical=50, overhead=80, profit=80, tax=20)
    payload_b = _make_payload("Estimate B", framing=650, roofing=40, electrical=75, overhead=120, profit=120, tax=25)

    comp = BidComp(llm_adapter=None)
    pair = comp._build_pair({"carrier": payload_a, "contractor": payload_b})  # type: ignore[arg-type]
    categories = comp._build_category_table(pair)
    framing_row = next(row for row in categories if row["category"] == "Framing / Structural")
    assert framing_row["primary_total"] == 400.0
    assert framing_row["comparison_total"] == 650.0

    overhead_row = next(row for row in categories if row["category"] == "Overhead & Profit")
    assert overhead_row["primary_total"] == 160.0
    assert overhead_row["comparison_total"] == 240.0

    top_deltas = comp._top_deltas(categories)
    narrative = comp._generate_narrative(pair, top_deltas)
    assert narrative.delta_rows
    assert narrative.blocks
    assert "Narrative fallback" in narrative.sections["overview_of_estimates"]
    assert narrative.key_drivers
    assert comp.last_narrative_debug["status"] == "fallback"
    assert "raw_response_preview" in comp.last_narrative_debug


def test_run_generates_two_tabs_with_llm() -> None:
    payload_a = _make_payload("Estimate A", framing=300, roofing=200, electrical=50, overhead=60, profit=60, tax=10)
    payload_b = _make_payload("Estimate B", framing=450, roofing=180, electrical=90, overhead=90, profit=90, tax=15)

    fake_markdown = """# Overview of Estimates
Estimate B carries higher framing scope.

## Key Cost Drivers
- **Framing / Structural**: 300 vs 450 — Estimate B replaces the entire deck.

## Scope Observations
- Estimate B includes extra electrical allowances.

## Suggested Follow-ups
- Confirm framing drawings for Estimate A.
"""
    sections = {
        "overview_of_estimates": "Estimate B carries higher framing scope.",
        "key_cost_drivers": [
            {
                "category": "Framing / Structural",
                "primary_total": 300,
                "comparison_total": 450,
                "delta_total": 150,
                "narrative": "Estimate B replaces the entire deck.",
            }
        ],
        "scope_observations": ["Estimate B includes extra electrical allowances."],
        "suggested_followups": ["Confirm framing drawings for Estimate A."],
    }
    comp = BidComp(llm_adapter=FakeAdapter(fake_markdown, sections))
    xlsx = comp.run({"carrier": payload_a, "contractor": payload_b}, job_id="job-1")
    wb = load_workbook(BytesIO(xlsx))
    assert wb.sheetnames == [
        "Summary",
        "Analysis",
    ]
    summary = wb["Summary"]
    assert summary["A1"].value == "Bid Comparison - Summary"
    assert summary["A9"].value == "Top Cost Drivers"
    assert comp.last_narrative_debug["status"] == "pipeline"


def test_comparison_name_falls_back_to_original_filename() -> None:
    payload_a = _make_payload("Carrier Estimate", framing=200, roofing=75, electrical=50, overhead=40, profit=40, tax=10)
    payload_b = _make_payload("", framing=220, roofing=60, electrical=55, overhead=44, profit=44, tax=11)
    payload_b["estimate_name"] = None
    payload_b["case_metadata"]["estimate_name"] = None
    payload_b["original_filename"] = "contractor_bid.pdf"

    comp = BidComp(llm_adapter=None)
    pair = comp._build_pair(
        {
            "carrier": payload_a,
            "contractor": {
                "payload": payload_b,
                "source_filename": "contractor_uploaded.pdf",
            },
        }
    )  # type: ignore[arg-type]

    assert pair.comparison.estimate_name == "contractor_bid.pdf"

