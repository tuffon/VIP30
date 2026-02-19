from __future__ import annotations

import json
from io import BytesIO

from openpyxl import load_workbook

from vip_shared.bid_comp import BidComp
from vip_shared.llm.adapter import LLMAdapterBase


def _make_payload(name: str, framing: float, roofing: float, electrical: float, overhead: float, profit: float, tax: float) -> dict:
    recap = {
        "O&P Items": [
            {"item": "FRAMING", "total": f"{framing:,.2f}"},
            {"item": "ROOFING", "total": f"{roofing:,.2f}"},
        ],
        "Non-O&P Items": [
            {"item": "ELECTRICAL", "total": f"{electrical:,.2f}"},
        ],
        "subtotals": [
            {"label": "Overhead", "total": f"{overhead:,.2f}"},
            {"label": "Profit", "total": f"{profit:,.2f}"},
            {"label": "Material Sales Tax", "total": f"{tax:,.2f}"},
            {"label": "Total", "total": f"{framing + roofing + electrical + overhead + profit + tax:,.2f}"},
        ],
    }
    return {
        "estimate_name": name,
        "case_metadata": {
            "estimate_name": name,
            "line_item_totals": {
                "grand_total": f"{framing + roofing + electrical + overhead + profit + tax:,.2f}",
                "material_sales_tax": f"{tax:,.2f}",
                "overhead_profit": f"{overhead + profit:,.2f}",
            },
        },
        "recaps_and_summaries": {"recap_by_category": recap},
        "sections": [],
    }


class FakeAdapter(LLMAdapterBase):
    def __init__(self, markdown: str, sections: dict | None = None) -> None:
        payload = {"markdown": markdown, "metadata": {"confidence": "high"}}
        if sections is not None:
            payload["sections"] = sections
        self.response = json.dumps(payload)

    def generate(self, template_id: str, context: dict) -> str:  # type: ignore[override]
        assert template_id == "bid_comp_summary_v1"
        assert "primary_json" in context
        assert "comparison_json" in context
        assert "category_table_json" in context
        return self.response


def test_category_mapping_and_fallback_narrative() -> None:
    payload_a = _make_payload("Estimate A", framing=400, roofing=100, electrical=50, overhead=80, profit=80, tax=20)
    payload_b = _make_payload("Estimate B", framing=650, roofing=40, electrical=75, overhead=120, profit=120, tax=25)

    comp = BidComp(llm_adapter=None)
    pair = comp._build_pair({"carrier": payload_a, "contractor": payload_b})  # type: ignore[arg-type]
    categories = comp._build_category_table(pair)
    framing_row = next(row for row in categories if row["category"] == "Framing / Structural")
    assert framing_row["primary_total"] == 400.0
    assert framing_row["comparison_total"] == 650.0

    overhead_row = next(row for row in categories if row["category"] == "Overhead & Profit")
    assert overhead_row["primary_total"] == 160.0
    assert overhead_row["comparison_total"] == 240.0

    top_deltas = comp._top_deltas(categories)
    narrative = comp._generate_narrative(pair, top_deltas)
    assert narrative.delta_rows
    assert narrative.blocks
    assert "Narrative fallback" in narrative.sections["overview_of_estimates"]
    assert narrative.key_drivers
    assert comp.last_narrative_debug["status"] == "fallback"
    assert "raw_response_preview" in comp.last_narrative_debug


def test_run_generates_two_tabs_with_llm() -> None:
    payload_a = _make_payload("Estimate A", framing=300, roofing=200, electrical=50, overhead=60, profit=60, tax=10)
    payload_b = _make_payload("Estimate B", framing=450, roofing=180, electrical=90, overhead=90, profit=90, tax=15)

    fake_markdown = """# Overview of Estimates
Estimate B carries higher framing scope.

## Key Cost Drivers
- **Framing / Structural**: 300 vs 450 — Estimate B replaces the entire deck.

## Scope Observations
- Estimate B includes extra electrical allowances.

## Suggested Follow-ups
- Confirm framing drawings for Estimate A.
"""
    sections = {
        "overview_of_estimates": "Estimate B carries higher framing scope.",
        "key_cost_drivers": [
            {
                "category": "Framing / Structural",
                "primary_total": 300,
                "comparison_total": 450,
                "delta_total": 150,
                "narrative": "Estimate B replaces the entire deck.",
            }
        ],
        "scope_observations": ["Estimate B includes extra electrical allowances."],
        "suggested_followups": ["Confirm framing drawings for Estimate A."],
    }
    comp = BidComp(llm_adapter=FakeAdapter(fake_markdown, sections))
    xlsx = comp.run({"carrier": payload_a, "contractor": payload_b}, job_id="job-1")
    wb = load_workbook(BytesIO(xlsx))
    assert wb.sheetnames == [
        "Summary",
        "Analysis",
    ]
    summary = wb["Summary"]
    assert summary["A1"].value == "Bid Comparison - Summary"
    assert summary["A9"].value == "Overall Summary"
    assert isinstance(summary["A10"].value, str)
    assert summary["A10"].value.strip() != ""
    assert summary["A12"].value == "Top Cost Drivers"
    assert comp.last_narrative_debug["status"] == "pipeline"


def test_summary_observations_redacts_internal_llm_errors() -> None:
    payload_a = _make_payload("Estimate A", framing=300, roofing=200, electrical=50, overhead=60, profit=60, tax=10)
    payload_b = _make_payload("Estimate B", framing=450, roofing=180, electrical=90, overhead=90, profit=90, tax=15)

    fake_markdown = """# Overview of Estimates
Analysis unavailable: Client error '429 Too Many Requests' for url 'https://api.openai.com/v1/chat/completions'

## Key Cost Drivers
- **Framing / Structural**: 300 vs 450

## Scope Observations
- Analysis unavailable: Client error '429 Too Many Requests' for url 'https://api.openai.com/v1/chat/completions'

## Suggested Follow-ups
- Retry.
"""
    sections = {
        "overview_of_estimates": "Analysis unavailable: Client error '429 Too Many Requests' for url 'https://api.openai.com/v1/chat/completions'",
        "key_cost_drivers": [
            {
                "category": "Framing / Structural",
                "primary_total": 300,
                "comparison_total": 450,
                "delta_total": 150,
                "narrative": "Analysis unavailable: Client error '429 Too Many Requests' for url 'https://api.openai.com/v1/chat/completions'",
            }
        ],
        "scope_observations": ["Analysis unavailable: Client error '429 Too Many Requests' for url 'https://api.openai.com/v1/chat/completions'"],
        "suggested_followups": ["Retry."],
    }

    comp = BidComp(llm_adapter=FakeAdapter(fake_markdown, sections))
    xlsx = comp.run({"carrier": payload_a, "contractor": payload_b}, job_id="job-2")
    wb = load_workbook(BytesIO(xlsx))
    summary = wb["Summary"]

    all_values = [
        str(cell.value)
        for row in summary.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    joined = "\n".join(all_values).lower()
    assert "too many requests" not in joined
    assert "chat/completions" not in joined
    assert (
        "narrative generation was temporarily rate-limited by the ai provider (http 429)." in joined
        or "automated narrative detail unavailable for this section." in joined
    )


def test_comparison_name_falls_back_to_original_filename() -> None:
    payload_a = _make_payload("Carrier Estimate", framing=200, roofing=75, electrical=50, overhead=40, profit=40, tax=10)
    payload_b = _make_payload("", framing=220, roofing=60, electrical=55, overhead=44, profit=44, tax=11)
    payload_b["estimate_name"] = None
    payload_b["case_metadata"]["estimate_name"] = None
    payload_b["original_filename"] = "contractor_bid.pdf"

    comp = BidComp(llm_adapter=None)
    pair = comp._build_pair(
        {
            "carrier": payload_a,
            "contractor": {
                "payload": payload_b,
                "source_filename": "contractor_uploaded.pdf",
            },
        }
    )  # type: ignore[arg-type]

    assert pair.comparison.estimate_name == "contractor_bid.pdf"



def test_map_category_reduces_unclassified_for_common_recap_labels() -> None:
    comp = BidComp(llm_adapter=None)

    expectations = {
        "LIGHT FIXTURES": "Electrical",
        "LABOR ONLY": "Miscellaneous / General Requirements",
        "SPECIALTY ITEMS": "Specialty Systems (low voltage, alarms, AV, solar)",
        "AWNINGS & PATIO COVERS": "Siding / Exterior Finishes",
        "ORNAMENTAL IRON": "Fencing / Gates",
        "FIREPLACES": "HVAC / Mechanical",
        "TILE": "Flooring",
        "FINISH HARDWARE": "Doors / Windows / Glass",
        "FIRE PROTECTION SYSTEMS": "Specialty Systems (low voltage, alarms, AV, solar)",
        "HAZARDOUS MATERIAL REMEDIATION": "Cleaning / Restoration",
    }

    for raw, expected in expectations.items():
        assert comp._map_category(raw) == expected


def test_aggregate_categories_maps_group_label_when_item_name_missing() -> None:
    comp = BidComp(llm_adapter=None)
    recap = {
        "Specialty Items": [{"total": "125.00"}],
        "Awnings & Patio Covers": [{"total": "400.00"}],
        "subtotals": [],
    }

    totals = comp._aggregate_categories(recap)
    assert totals["Specialty Systems (low voltage, alarms, AV, solar)"] == 125.0
    assert totals["Siding / Exterior Finishes"] == 400.0
    assert totals["Other / Unclassified"] == 0.0
